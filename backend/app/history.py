from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from .database import DetectedSetup, MLSignalSnapshot, SetupObservation, TradePlanVariant
from .models import (
    HistoricalPlanView,
    HistoryItem,
    HistoryResponse,
    ScanResult,
    SetupDetailResponse,
    TradePlanResultView,
)


def iso(value: Optional[datetime]) -> Optional[str]:
    if value is None:
        return None
    normalized = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    return normalized.isoformat()


def history_item(setup: DetectedSetup) -> HistoryItem:
    latest = setup.observations[-1] if setup.observations else None
    return HistoryItem(
        id=setup.id,
        symbol=setup.symbol,
        direction=setup.direction,
        first_seen_at=iso(setup.first_seen_at) or "",
        last_seen_at=iso(setup.last_seen_at) or "",
        rating=latest.rating if latest else 0,
        setup_class=latest.setup_class if latest else "Weak",
        support_level=setup.support_level,
        resistance_level=setup.resistance_level,
        entry_price=setup.entry_price,
        stop_loss=setup.stop_loss,
        take_profit=setup.take_profit,
        reward_risk=setup.reward_risk,
        trade_plan_status=setup.trade_plan_status,
        outcome=setup.outcome,
        entered_at=iso(setup.entered_at),
        resolved_at=iso(setup.resolved_at),
        price_at_deadline=setup.price_at_deadline,
        mfe_r=setup.mfe_r,
        mae_r=setup.mae_r,
        ambiguous_intrabar=setup.ambiguous_intrabar,
    )


def _filters(
    *,
    symbol: Optional[str],
    direction: Optional[str],
    outcome: Optional[str],
    min_rating: Optional[int],
    date_from: Optional[datetime],
    date_to: Optional[datetime],
) -> list:
    clauses = []
    if symbol:
        clauses.append(DetectedSetup.symbol.ilike(f"%{symbol.strip()}%"))
    if direction:
        clauses.append(DetectedSetup.direction == direction)
    if outcome:
        clauses.append(DetectedSetup.outcome == outcome)
    if date_from:
        clauses.append(DetectedSetup.first_seen_at >= date_from)
    if date_to:
        clauses.append(DetectedSetup.first_seen_at <= date_to)
    if min_rating is not None:
        latest_rating = (
            select(SetupObservation.rating)
            .where(SetupObservation.setup_id == DetectedSetup.id)
            .order_by(SetupObservation.observed_at.desc())
            .limit(1)
            .scalar_subquery()
        )
        clauses.append(latest_rating >= min_rating)
    return clauses


def query_history(
    session: Session,
    *,
    symbol: Optional[str] = None,
    direction: Optional[str] = None,
    outcome: Optional[str] = None,
    min_rating: Optional[int] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    page: int = 1,
    page_size: int = 50,
) -> HistoryResponse:
    clauses = _filters(
        symbol=symbol,
        direction=direction,
        outcome=outcome,
        min_rating=min_rating,
        date_from=date_from,
        date_to=date_to,
    )
    count_stmt = select(func.count(DetectedSetup.id)).where(*clauses)
    total = session.scalar(count_stmt) or 0
    stmt = (
        select(DetectedSetup)
        .options(selectinload(DetectedSetup.observations))
        .where(*clauses)
        .order_by(DetectedSetup.first_seen_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    setups = session.scalars(stmt).all()
    items = [history_item(setup) for setup in setups]
    return HistoryResponse(page=page, page_size=page_size, total=total, items=items)


def query_setup_detail(session: Session, setup_id: int) -> Optional[SetupDetailResponse]:
    setup = session.scalar(
        select(DetectedSetup)
        .options(selectinload(DetectedSetup.observations))
        .where(DetectedSetup.id == setup_id)
    )
    if setup is None or not setup.observations:
        return None
    observation = setup.observations[0]
    result = ScanResult.model_validate(observation.result_json)
    snapshot = session.scalar(
        select(MLSignalSnapshot)
        .options(
            selectinload(MLSignalSnapshot.plan_variants).selectinload(TradePlanVariant.results)
        )
        .where(MLSignalSnapshot.setup_id == setup_id)
        .order_by(MLSignalSnapshot.observed_at)
        .limit(1)
    )
    plans: list[HistoricalPlanView] = []
    if snapshot is not None:
        for plan in snapshot.plan_variants:
            plans.append(
                HistoricalPlanView(
                    id=plan.id,
                    version=plan.plan_version,
                    status=plan.status,
                    reason=plan.reason,
                    direction=plan.direction,
                    activation=plan.activation,
                    entry_price=plan.entry_price,
                    stop_loss=plan.stop_loss,
                    risk_price=plan.risk_price,
                    target_1r=plan.target_1r,
                    target_2r=plan.target_2r,
                    target_3r=plan.target_3r,
                    trigger_price=plan.trigger_price,
                    retest_zone_low=plan.retest_zone_low,
                    retest_zone_high=plan.retest_zone_high,
                    results=[
                        TradePlanResultView(
                            horizon_minutes=item.horizon_minutes,
                            outcome=item.outcome,
                            entry_price=item.entry_price,
                            stop_loss=item.stop_loss,
                            target_1r=item.target_1r,
                            target_2r=item.target_2r,
                            target_3r=item.target_3r,
                            entered_at=iso(item.entered_at),
                            stopped_at=iso(item.stopped_at),
                            hit_1r_at=iso(item.hit_1r_at),
                            hit_2r_at=iso(item.hit_2r_at),
                            hit_3r_at=iso(item.hit_3r_at),
                            mfe_r=item.mfe_r,
                            mae_r=item.mae_r,
                            ambiguous_intrabar=item.ambiguous_intrabar,
                        )
                        for item in sorted(plan.results, key=lambda value: value.horizon_minutes)
                    ],
                )
            )
        result.trade_plan_variants = plans
    return SetupDetailResponse(
        setup=history_item(setup),
        detector_version=snapshot.detector_version if snapshot is not None else "v1",
        feature_schema_version=snapshot.feature_schema_version if snapshot is not None else "legacy",
        snapshot_id=snapshot.id if snapshot is not None else None,
        result=result,
        plans=plans,
    )


def history_candle_rows(session: Session, setup_ids: list[int]) -> list[list]:
    if not setup_ids:
        return []
    rows: list[list] = []
    observations = session.execute(
        select(SetupObservation, DetectedSetup.symbol)
        .join(DetectedSetup, DetectedSetup.id == SetupObservation.setup_id)
        .where(SetupObservation.setup_id.in_(setup_ids))
        .order_by(SetupObservation.observed_at, SetupObservation.setup_id)
    ).all()
    for observation, symbol in observations:
        for candle in observation.range_candles_json:
            rows.append(
                [
                    observation.setup_id,
                    symbol,
                    iso(observation.observed_at),
                    candle.get("timestamp"),
                    candle.get("open"),
                    candle.get("high"),
                    candle.get("low"),
                    candle.get("close"),
                    candle.get("volume"),
                    candle.get("turnover"),
                ]
            )
    return rows


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1D2633")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_index, column in enumerate(sheet.columns, start=1):
        longest = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(longest + 2, 10), 28)


def export_history_xlsx(response: HistoryResponse, candle_rows: Optional[list[list]] = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "История"
    headers = [
        "ID",
        "Монета",
        "Направление",
        "Первое обнаружение UTC",
        "Последнее обнаружение UTC",
        "Рейтинг",
        "Класс",
        "Support",
        "Resistance",
        "ТВХ",
        "Стоп",
        "Тейк",
        "RR",
        "Статус плана",
        "Исход",
        "Вход UTC",
        "Завершение UTC",
        "Цена через час",
        "MFE, R",
        "MAE, R",
        "Неоднозначная M1",
    ]
    sheet.append(headers)
    for item in response.items:
        sheet.append(
            [
                item.id,
                item.symbol,
                item.direction,
                item.first_seen_at,
                item.last_seen_at,
                item.rating,
                item.setup_class,
                item.support_level,
                item.resistance_level,
                item.entry_price,
                item.stop_loss,
                item.take_profit,
                item.reward_risk,
                item.trade_plan_status,
                item.outcome,
                item.entered_at,
                item.resolved_at,
                item.price_at_deadline,
                item.mfe_r,
                item.mae_r,
                "Да" if item.ambiguous_intrabar else "Нет",
            ]
        )

    _style_sheet(sheet)

    candles_sheet = workbook.create_sheet("Свечи диапазона")
    candles_sheet.append(
        ["Setup ID", "Монета", "Наблюдение UTC", "Timestamp ms", "Open", "High", "Low", "Close", "Volume", "Turnover"]
    )
    for row in candle_rows or []:
        candles_sheet.append(row)
    _style_sheet(candles_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
